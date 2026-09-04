"""detect.py tested against the real sample workbooks plus synthetic edge
cases (empty workbook, unrelated sheet names, a corrupt/non-xlsx file) --
content-based classification must never rely on filename.
"""
from pathlib import Path

import openpyxl
import pytest

from src.pipeline.detect import detect_workbook_type

RAW_DIR = Path(__file__).resolve().parents[1] / "data" / "raw"
DEC_FILE = RAW_DIR / "Dec 2025 Coverage Analysis (0-11).xlsx"
VPD_FILE = RAW_DIR / "KP VPDs Line List Week 1-32,2026.xlsx"


@pytest.mark.skipif(not DEC_FILE.exists(), reason="Real raw Excel file not present in this environment")
def test_detects_coverage_workbook_by_content_even_with_unrelated_filename(tmp_path):
    # Rename to something with no "coverage" hint at all -- detection must
    # still work, since it was never supposed to depend on the filename.
    renamed = tmp_path / "randomly_named_file_1.xlsx"
    renamed.write_bytes(DEC_FILE.read_bytes())
    result = detect_workbook_type(renamed)
    assert result.workbook_type == "coverage"
    assert len(result.matched_sheets) >= 2


@pytest.mark.skipif(not VPD_FILE.exists(), reason="Real raw Excel file not present in this environment")
def test_detects_vpd_workbook_by_content_even_without_vpd_in_filename(tmp_path):
    # This is the exact bug the old filename-sniffing approach had: a real
    # VPD line list not named with "VPD" in it would have been silently
    # treated as a coverage file (or ignored) before this module existed.
    renamed = tmp_path / "weekly_disease_report.xlsx"
    renamed.write_bytes(VPD_FILE.read_bytes())
    result = detect_workbook_type(renamed)
    assert result.workbook_type == "vpd"
    assert len(result.matched_sheets) >= 2


def test_unrelated_sheet_names_are_unknown_not_guessed(tmp_path):
    path = tmp_path / "some_export.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    wb.active["A1"] = "hello"
    wb.create_sheet("Sheet2")
    wb.save(path)
    result = detect_workbook_type(path)
    assert result.workbook_type == "unknown"
    assert "Monitoring" in result.message  # explains why, doesn't fabricate a match


def test_empty_workbook_is_flagged_not_crashed_on(tmp_path):
    path = tmp_path / "blank.xlsx"
    wb = openpyxl.Workbook()
    wb.save(path)  # a fresh Workbook always has one sheet ("Sheet") -- still not coverage/VPD
    result = detect_workbook_type(path)
    assert result.workbook_type == "unknown"


def test_unreadable_file_is_flagged_not_crashed_on(tmp_path):
    path = tmp_path / "not_really_excel.xlsx"
    path.write_text("this is plain text, not a real xlsx file")
    result = detect_workbook_type(path)
    assert result.workbook_type == "unreadable"
    assert "could not be opened" in result.message


def test_one_matching_sheet_is_not_enough_to_confidently_classify(tmp_path):
    # Only one sheet name overlaps with the coverage signature -- below
    # MIN_MATCHING_SHEETS, so this must not be misclassified as coverage.
    path = tmp_path / "one_sheet_overlap.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "District "
    wb.create_sheet("Something Else Entirely")
    wb.save(path)
    result = detect_workbook_type(path)
    assert result.workbook_type == "unknown"
