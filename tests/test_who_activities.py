"""WHO Supported Activities tests, pinned to real numbers confirmed by
direct inspection of WHO_EPI_May_2024_Highlights_Dashboard.xlsx (a single
duty station's May 2024 field-monitoring highlights report, not a
multi-period activity database -- see who_activities.py's module
docstring), same convention as the other real-data test modules.
"""
from pathlib import Path

import pytest

from src.pipeline.config import DISTRICT_TO_BOUNDARY
from src.pipeline.detect import detect_workbook_type
from src.pipeline.who_activities import (
    ACTIVITY_THEME, build_summary, clean_evidence_table, find_who_activities_files, load_who_activities,
)

RAW_DIR = Path(__file__).resolve().parents[1] / "data" / "raw"
WHO_FILE = RAW_DIR / "WHO_EPI_May_2024_Highlights_Dashboard.xlsx"

pytestmark = pytest.mark.skipif(
    not WHO_FILE.exists(),
    reason="Real WHO_EPI_May_2024_Highlights_Dashboard.xlsx not present in this environment",
)


@pytest.fixture(scope="module")
def loaded():
    return load_who_activities(WHO_FILE)


@pytest.fixture(scope="module")
def summary(loaded):
    return build_summary(loaded)


def test_detects_as_who_activities_not_another_domain():
    result = detect_workbook_type(WHO_FILE)
    assert result.workbook_type == "who_activities"


def test_find_who_activities_files_locates_it_in_raw_dir():
    assert WHO_FILE in find_who_activities_files(RAW_DIR)


def test_headline_kpis_match_source(loaded):
    kpis = loaded["kpis"]
    assert kpis["field_support_days"] == "11"
    assert kpis["zero_dose_vaccinated_arandu"] == "221"
    assert kpis["measles_results_reviewed"] == "8"
    assert kpis["mobr_districts_supported"] == "3"


def test_evidence_table_has_11_rows_no_duplicates(loaded):
    df = loaded["evidence_df"]
    assert len(df) == 11
    assert df["id"].duplicated().sum() == 0
    assert df.isna().sum().sum() == 0


def test_who_contributions_stops_before_next_section_header(loaded):
    # Regression test: column H on the source sheet runs contribution
    # bullets directly into the next section's own header with no blank
    # separator row -- confirmed real bug found and fixed this session.
    assert len(loaded["who_contributions"]) == 6
    for c in loaded["who_contributions"]:
        assert not c.lstrip("• ").startswith(("4.", "PRIORITY"))


def test_key_results_and_district_highlights_counts(loaded):
    assert len(loaded["key_results"]) == 5
    assert len(loaded["district_highlights"]) == 6
    assert len(loaded["priority_gaps"]) == 6


def test_district_name_variants_are_standardized(loaded):
    df, unmapped = clean_evidence_table(loaded["evidence_df"])
    assert unmapped == []
    # Source spells these adjective-first; this project's canonical spelling
    # (config.DISTRICT_TO_BOUNDARY) is place-name-first.
    assert "Lower Chitral" not in df["district_canonical"].values
    assert "Upper Chitral" not in df["district_canonical"].values
    assert "Chitral Lower" in df["district_canonical"].values
    assert "Chitral Upper" in df["district_canonical"].values
    real_districts = set(df["district_canonical"]) - {"Multiple districts"}
    assert real_districts <= set(DISTRICT_TO_BOUNDARY)


def test_every_evidence_row_has_an_activity_theme(loaded):
    df, _ = clean_evidence_table(loaded["evidence_df"])
    assert df["activity_theme"].isna().sum() == 0
    assert set(df["id"]) == set(ACTIVITY_THEME)


def test_activity_theme_breakdown_sums_to_total_records(summary):
    assert sum(summary["activity_theme_breakdown"].values()) == 11


def test_district_breakdown_sums_to_total_records(summary):
    assert sum(r["evidence_count"] for r in summary["district_breakdown"]) == 11


def test_district_map_excludes_multiple_districts_and_has_no_unmapped(summary):
    dmap = summary["district_map"]
    assert dmap["unmapped_districts"] == []
    assert "Multiple districts" not in dmap["features"]
    assert dmap["features"]["Dir Lower"]["evidence_count"] == 4
    assert dmap["features"]["Chitral Lower"]["evidence_count"] == 3
    assert dmap["features"]["Malakand"]["evidence_count"] == 1
    assert dmap["features"]["Dir Upper"]["evidence_count"] == 1
    assert dmap["features"]["Chitral Upper"]["evidence_count"] == 1
    # Only this duty station's 6 assigned districts should ever appear --
    # every other KP district is genuinely outside this report's scope.
    assert len(dmap["features"]) == 6
    assert dmap["features"]["Bajaur"]["evidence_count"] == 0


def test_bajaur_has_narrative_highlight_but_no_evidence_row(summary):
    # A real, documented gap in the source data -- Bajaur appears in the
    # Highlights sheet's district-level highlights but has no row of its own
    # in the Evidence & Findings table. Surfaced (and shown as a real 0, not
    # omitted), not silently smoothed over.
    assert "Bajaur" in summary["data_quality"]["districts_with_narrative_highlight_but_no_evidence_row"]
    assert summary["district_map"]["features"]["Bajaur"]["evidence_count"] == 0


def test_timeline_excludes_month_only_rows_and_is_chronological(summary):
    timeline = summary["timeline"]
    assert len(timeline) == 9  # 11 evidence rows minus 2 with no specific day (E01, E05)
    dates = [t["date"] for t in timeline]
    assert dates == sorted(dates)
    assert timeline[0]["date"] == "2024-05-08"
    assert timeline[-1]["date"] == "2024-05-30"
    range_row = next(t for t in timeline if t["id"] == "E08")
    assert range_row["date"] == "2024-05-20"
    assert range_row["date_end"] == "2024-05-25"


def test_measles_review_pct_uses_the_source_stated_denominator(summary):
    assert summary["kpis"]["measles_results_total"] == 11
    assert summary["kpis"]["measles_review_pct"] == pytest.approx(8 / 11 * 100, abs=0.1)


def test_headline_kpis_reconcile_against_raw_narrative_and_summary_table(summary):
    checks = summary["data_quality"]["headline_kpis_reconciled_against_summary_table_and_raw_narrative"]
    assert len(checks) == 4
    for c in checks:
        assert c["found_in_raw_narrative"] is True


def test_evidence_records_json_serializes_without_bare_nan(summary):
    # Regression test: assigning a list containing None back into an
    # existing datetime64 DataFrame column silently re-coerces None to NaN
    # (a real pandas quirk found and fixed this session) -- json.dump still
    # writes a bare `NaN` token for that (Python's json module tolerates it,
    # but it is not valid JSON and breaks JSON.parse() in the browser).
    import json
    text = json.dumps(summary, default=str)
    assert "NaN" not in text
    month_only_rows = [r for r in summary["evidence_records"] if r["is_month_only"]]
    assert len(month_only_rows) == 2
    for r in month_only_rows:
        assert r["date_start"] is None
        assert r["date_end"] is None


def test_excel_export_includes_who_sheets_with_correct_row_counts(tmp_path, summary):
    from src.pipeline.export_excel import build_processed_excel
    from src.pipeline.run_vpd import PROCESSED_DIR
    import openpyxl

    out = build_processed_excel(PROCESSED_DIR, tmp_path / "export.xlsx")
    assert out is not None
    wb = openpyxl.load_workbook(out)
    for name in ["WHO Activities Summary", "WHO Evidence Data", "WHO District Summary",
                 "WHO Activity Theme Summary", "WHO Data Quality"]:
        assert name in wb.sheetnames

    evidence_ws = wb["WHO Evidence Data"]
    assert evidence_ws.max_row - 1 == 11  # header + 11 records
    assert evidence_ws.freeze_panes == "A2"

    district_ws = wb["WHO District Summary"]
    assert district_ws.max_row - 1 == len(summary["district_map"]["features"])


def test_no_status_or_target_achievement_fields_are_fabricated(summary):
    # This workbook has no activity-status column and no target/achievement
    # pair beyond the measles-review KPI -- confirmed by inspection, so
    # neither should appear anywhere in the summary.
    assert "status_breakdown" not in summary
    assert "target_achievement" not in summary
    for r in summary["evidence_records"]:
        assert "status" not in r
