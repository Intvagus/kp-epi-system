"""Multi-sheet Excel export of whatever processed data exists for a run.

Coverage and VPD are each independent -- a sheet only gets added for a
component that actually processed, never an empty or broken one for a
component that wasn't uploaded. Reads the SAME data/processed/* files the
dashboard reads and never recomputes anything, so the two can never disagree
(same rule as the bulletin -- see src/bulletin/build.py).

"""
import json
from pathlib import Path

import pandas as pd


def _coverage_district_frame(processed_dir: Path) -> pd.DataFrame | None:
    path = processed_dir / "coverage_district.parquet"
    return pd.read_parquet(path) if path.exists() else None


def _coverage_uc_frame(processed_dir: Path) -> pd.DataFrame | None:
    path = processed_dir / "coverage_uc.parquet"
    return pd.read_parquet(path) if path.exists() else None


def _coverage_summary(processed_dir: Path) -> dict | None:
    path = processed_dir / "coverage_summary.json"
    if not path.exists():
        return None
    with open(path, encoding="utf-8") as f:
        summary = json.load(f)
    return summary if summary.get("status") == "ok" else None


def _coverage_kpi_rows(summary: dict) -> list[dict]:
    rows = []
    for kind, period in summary["periods"].items():
        if period.get("status") != "ok":
            continue
        exe = period["executive"]
        dc = exe["district_compliance"]
        rows.append({
            "period_kind": kind, "period_label": exe["current_period_label"],
            "fic_pct": exe["fic_pct"], "dropout_pct": exe["dropout_pct"],
            "target_surviving_infants": exe["target_surviving_infants"],
            "best_antigen": exe["best_antigen"]["antigen"] if exe["best_antigen"] else None,
            "best_antigen_pct": exe["best_antigen"]["pct"] if exe["best_antigen"] else None,
            "worst_antigen": exe["worst_antigen"]["antigen"] if exe["worst_antigen"] else None,
            "worst_antigen_pct": exe["worst_antigen"]["pct"] if exe["worst_antigen"] else None,
            "best_district": exe["best_district"]["district"] if exe["best_district"] else None,
            "best_district_pct": exe["best_district"]["pct"] if exe["best_district"] else None,
            "worst_district": exe["worst_district"]["district"] if exe["worst_district"] else None,
            "worst_district_pct": exe["worst_district"]["pct"] if exe["worst_district"] else None,
            "districts_achieving_target": dc["good"],
            "districts_requiring_intervention": dc["warning"] + dc["poor"],
            "districts_total": dc["total_with_data"],
        })
    return rows


def _coverage_antigen_rows(summary: dict) -> list[dict]:
    rows = []
    for kind, period in summary["periods"].items():
        if period.get("status") != "ok":
            continue
        for a in period["antigen_analysis"]:
            rows.append({"period_kind": kind, **a})
    return rows


def _vpd_summary(processed_dir: Path) -> dict | None:
    path = processed_dir / "vpd_summary.json"
    if not path.exists():
        return None
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def _vpd_summary_rows(vpd: dict) -> list[dict]:
    return [
        {"metric": "Latest epi week", "value": vpd.get("latest_epi_week")},
        {"metric": "MSL suspected (YTD)", "value": vpd["msl"]["suspected_ytd"]},
        {"metric": "MSL suspected (latest week)", "value": vpd["msl"]["suspected_latest_week"]},
        {"metric": "MSL sample collection rate %", "value": vpd["msl"]["sample_collection_rate_pct"]},
        {"metric": "MSL sample adequacy rate %", "value": vpd["msl"]["sample_adequacy_rate_pct"]},
        {"metric": "Diphtheria cases (YTD)", "value": vpd["diphtheria"]["ytd"]["case_count"]},
        {"metric": "Diphtheria districts affected", "value": vpd["diphtheria"]["ytd"]["districts_affected"]},
        {"metric": "Diphtheria lab confirmed", "value": vpd["diphtheria"]["ytd"]["lab_confirmed_count"]},
        {"metric": "Diphtheria deaths", "value": vpd["diphtheria"]["ytd"]["deaths"]},
        {"metric": "NNT cases (YTD)", "value": vpd["nnt"]["ytd"]["case_count"]},
        {"metric": "NNT deaths", "value": vpd["nnt"]["ytd"]["deaths"]},
    ]


def _vpd_district_rows(vpd: dict) -> list[dict]:
    rows = []
    for r in vpd["msl"]["district_breakdown"]:
        rows.append({"disease": "Measles-Rubella", **r})
    for r in vpd["pertussis"]["district_breakdown"]:
        rows.append({"disease": "Pertussis", **r})
    for r in vpd["nnt"]["by_district_ytd"]:
        rows.append({"disease": "NNT", **r})
    return rows


def _who_activities_summary(processed_dir: Path) -> dict | None:
    path = processed_dir / "who_activities_summary.json"
    if not path.exists():
        return None
    with open(path, encoding="utf-8") as f:
        summary = json.load(f)
    return summary if summary.get("status") == "ok" else None


def _who_summary_rows(w: dict) -> list[dict]:
    k = w["kpis"]
    return [
        {"metric": "Reporting period", "value": w["reporting_period_label"]},
        {"metric": "Duty station", "value": w["duty_station"]},
        {"metric": "Field-support days/visits", "value": k["field_support_days"]},
        {"metric": "Zero-dose children vaccinated (Arandu)", "value": k["zero_dose_vaccinated_arandu"]},
        {"metric": "Measles case results reviewed", "value": k["measles_results_reviewed"]},
        {"metric": "Measles case results total (per source)", "value": k["measles_results_total"]},
        {"metric": "Measles case-result review %", "value": k["measles_review_pct"]},
        {"metric": "MOBR-affected districts supported", "value": k["mobr_districts_supported"]},
        {"metric": "Districts covered", "value": k["districts_covered"]},
        {"metric": "Evidence & Findings records", "value": k["evidence_record_count"]},
    ]


def _who_evidence_rows(w: dict) -> list[dict]:
    cols = ["id", "district_canonical", "date_period_raw", "date_start", "date_end", "activity_theme",
            "activity", "quantified_result", "who_contribution", "suggested_highlight", "evidence_source"]
    rename = {
        "id": "ID", "district_canonical": "District", "date_period_raw": "Date / Period (source)",
        "date_start": "Date (parsed start)", "date_end": "Date (parsed end)", "activity_theme": "Activity Theme",
        "activity": "Activity / Finding", "quantified_result": "Quantified Result",
        "who_contribution": "WHO Contribution / Action", "suggested_highlight": "Suggested Highlight",
        "evidence_source": "Evidence Source",
    }
    return [{rename[c]: r.get(c) for c in cols} for r in w["evidence_records"]]


def _who_district_rows(w: dict) -> list[dict]:
    features = w["district_map"]["features"]
    return sorted(
        ({"district": d, "evidence_count": v["evidence_count"]} for d, v in features.items()),
        key=lambda r: -r["evidence_count"],
    )


def _who_theme_rows(w: dict) -> list[dict]:
    total = sum(w["activity_theme_breakdown"].values())
    return [
        {"activity_theme": t, "evidence_count": n, "pct_of_records": round(n / total * 100, 1) if total else None}
        for t, n in sorted(w["activity_theme_breakdown"].items(), key=lambda kv: -kv[1])
    ]


def _who_data_quality_rows(w: dict) -> list[dict]:
    dq = w["data_quality"]
    rows = [
        {"check": "Evidence & Findings record count", "result": dq["evidence_record_count"]},
        {"check": "Duplicate evidence IDs", "result": dq["duplicate_evidence_ids"]},
        {"check": "Unmapped district names", "result": ", ".join(dq["unmapped_districts"]) or "None"},
        {"check": "District name variants standardized", "result": "; ".join(
            f"{k} -> {v}" for k, v in dq["district_name_variants_standardized"].items()) or "None"},
        {"check": "Districts with narrative highlight but no evidence row", "result": ", ".join(
            dq["districts_with_narrative_highlight_but_no_evidence_row"]) or "None"},
        {"check": "Districts with evidence row but no narrative highlight", "result": ", ".join(
            dq["districts_with_evidence_row_but_no_narrative_highlight"]) or "None"},
    ]
    for col, n in dq["missing_values_by_column"].items():
        rows.append({"check": f"Missing values -- {col}", "result": n})
    for c in dq["headline_kpis_reconciled_against_summary_table_and_raw_narrative"]:
        rows.append({
            "check": f"Reconciled: {c['metric']}",
            "result": f"{c['headline_value']} (found in raw narrative report: {c['found_in_raw_narrative']})",
        })
    return rows


def _monitoring_summary(processed_dir: Path) -> dict | None:
    path = processed_dir / "monitoring_summary.json"
    if not path.exists():
        return None
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def _rca_summary_rows(rca: dict) -> list[dict]:
    ov = rca["overview"]
    zd = rca["zero_dose"]
    return [
        {"metric": "Reporting window", "value": rca["reporting_window"]},
        {"metric": "Total children assessed", "value": ov["total_children_assessed"]},
        {"metric": "Total RCA visits", "value": ov["total_rca_visits"]},
        {"metric": "Districts covered", "value": ov["districts_covered"]},
        {"metric": "Zero-dose children (Penta1 not received)", "value": zd["zero_dose_count"]},
        {"metric": "Zero-dose %", "value": zd["zero_dose_pct"]},
    ]


def _supervisory_summary_rows(sup: dict) -> list[dict]:
    ov = sup["overview"]
    rows = [
        {"metric": "Reporting window", "value": sup["reporting_window"]},
        {"metric": "Total supervisory visits", "value": ov["total_visits"]},
        {"metric": "Districts covered", "value": ov["districts_covered"]},
        {"metric": "Facilities covered", "value": ov["facilities_covered"]},
        {"metric": "Fixed-site-open rate %", "value": sup["fixed_site_open_rate"]["pct"]},
    ]
    for s in sup["composite_scores"]:
        rows.append({"metric": f"{s['category']} (avg %)", "value": s["avg_pct"]})
    return rows


def build_processed_excel(processed_dir: Path, output_path: Path) -> Path | None:
    """Writes one .xlsx with a sheet per available, actually-processed
    component. Returns None (writes nothing) if no processed data exists at
    all -- there being nothing to export yet is not an error."""
    sheets: dict[str, pd.DataFrame] = {}

    district_df = _coverage_district_frame(processed_dir)
    if district_df is not None and not district_df.empty:
        sheets["Coverage District Data"] = district_df
        uc_df = _coverage_uc_frame(processed_dir)
        if uc_df is not None and not uc_df.empty:
            sheets["Coverage UC Data"] = uc_df
        coverage_summary = _coverage_summary(processed_dir)
        if coverage_summary:
            kpi_rows = _coverage_kpi_rows(coverage_summary)
            if kpi_rows:
                sheets["Coverage KPI Summary"] = pd.DataFrame(kpi_rows)
            antigen_rows = _coverage_antigen_rows(coverage_summary)
            if antigen_rows:
                sheets["Coverage Antigen Summary"] = pd.DataFrame(antigen_rows)

    vpd = _vpd_summary(processed_dir)
    if vpd:
        sheets["VPD Surveillance Summary"] = pd.DataFrame(_vpd_summary_rows(vpd))
        district_rows = _vpd_district_rows(vpd)
        if district_rows:
            sheets["VPD District Breakdown"] = pd.DataFrame(district_rows)

    monitoring = _monitoring_summary(processed_dir)
    if monitoring:
        rca = monitoring.get("rca", {})
        if rca.get("status") == "ok":
            sheets["RCA Summary"] = pd.DataFrame(_rca_summary_rows(rca))
            sheets["RCA Antigen Coverage"] = pd.DataFrame(rca["antigen_coverage"])
            if rca["district_breakdown"]:
                sheets["RCA District Breakdown"] = pd.DataFrame(rca["district_breakdown"])
        sup = monitoring.get("supervisory", {})
        if sup.get("status") == "ok":
            sheets["Supervisory Summary"] = pd.DataFrame(_supervisory_summary_rows(sup))
            sheets["Supervisory Compliance Items"] = pd.DataFrame(sup["compliance_items"])
            if sup["district_breakdown"]:
                sheets["Supervisory District Breakdown"] = pd.DataFrame(sup["district_breakdown"])

    who = _who_activities_summary(processed_dir)
    who_sheet_names = set()
    if who:
        sheets["WHO Activities Summary"] = pd.DataFrame(_who_summary_rows(who))
        sheets["WHO Evidence Data"] = pd.DataFrame(_who_evidence_rows(who))
        sheets["WHO District Summary"] = pd.DataFrame(_who_district_rows(who))
        sheets["WHO Activity Theme Summary"] = pd.DataFrame(_who_theme_rows(who))
        sheets["WHO Data Quality"] = pd.DataFrame(_who_data_quality_rows(who))
        who_sheet_names = {"WHO Activities Summary", "WHO Evidence Data", "WHO District Summary",
                            "WHO Activity Theme Summary", "WHO Data Quality"}

    if not sheets:
        return None

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name[:31], index=False)  # Excel's 31-char sheet-name limit
            if name in who_sheet_names:
                _format_who_sheet(writer.sheets[name[:31]], df)
    return output_path


def _format_who_sheet(ws, df: pd.DataFrame) -> None:
    """Professional formatting for the WHO Supported Activities sheets:
    bold header row, frozen header + (for the row-per-record sheets) frozen
    ID/District columns, a header-row filter, and column widths sized to
    content rather than Excel's default. Scoped to just these sheets since
    that's what this export was asked to be "professionally structured" --
    not a general reformat of every sheet this exporter already writes."""
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A3A7A", end_color="1A3A7A", fill_type="solid")  # dashboard's navy
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for i, col in enumerate(df.columns, start=1):
        sample_lengths = [len(str(col))] + [len(str(v)) for v in df[col].astype(str).head(200)]
        width = min(max(sample_lengths) + 2, 60)
        ws.column_dimensions[get_column_letter(i)].width = width
        if "pct" in col.lower() or "%" in col:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=i)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.0"
