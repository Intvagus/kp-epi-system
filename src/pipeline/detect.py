"""Content-based classification of an uploaded .xlsx workbook: Coverage,
VPD surveillance, or unrecognized -- by comparing its actual sheet names
against the two known signatures, never by filename convention.

Filename sniffing (the old approach: "vpd" in the filename means VPD, else
assume coverage) breaks the moment a real upload isn't named the way this
project's own sample files happen to be named. Sheet names are the one thing
both source files actually guarantee, per CLAUDE.md's confirmed structure.

Monitoring/supervisory-visit data has no known signature here on purpose --
no sample file has ever been received (see CLAUDE.md's open questions), so
there is nothing to detect against. A workbook that isn't a Coverage or VPD
match falls out as "unknown" with an honest message, never guessed at.
"""
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
import pandas as pd

from .admin_activities import ADMIN_ACTIVITIES_SHEET, TASK_COLUMN_HEADER as ADMIN_ACTIVITIES_HEADER
from .config import SHEET_NAMES, VPD_SHEET_NAMES
from .indicator_sheet_vpd import INDICATOR_SHEET_TITLE_MARKER
from .who_activities import REQUIRED_SHEETS as WHO_ACTIVITIES_SIGNATURE

COVERAGE_SIGNATURE = {name.strip().lower() for name in SHEET_NAMES.values()}
VPD_SIGNATURE = {name.strip().lower() for name in VPD_SHEET_NAMES.values()}
WHO_ACTIVITIES_SIGNATURE = {name.strip().lower() for name in WHO_ACTIVITIES_SIGNATURE}

# A workbook doesn't need every expected sheet to match (e.g. a re-export
# missing one sheet) -- this many of the signature's sheet names present is
# enough to call it confidently, while still ruling out a coincidental
# single-sheet-name overlap with an unrelated file.
MIN_MATCHING_SHEETS = 2


@dataclass
class DetectionResult:
    workbook_type: str  # "coverage" | "vpd" | "indicator_sheet" | "who_activities" | "admin_activities" | "unknown" | "empty" | "unreadable"
    matched_sheets: list = field(default_factory=list)
    all_sheets: list = field(default_factory=list)
    message: str = ""


def _sheet_names(path: Path) -> list[str]:
    return pd.ExcelFile(path, engine="openpyxl").sheet_names


def _is_indicator_sheet_workbook(path: Path) -> bool:
    """The Measles Indicator Sheet workbook is a completely different shape
    from Coverage/VPD (one sheet per year, e.g. '2026', not a fixed set of
    named sheets) so it can't use the sheet-name-signature approach those
    two use. Its one reliable, content-based marker: cell A1 of at least one
    sheet literally reads '...Indicator Sheet-<year>' in the source file."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return False
    for name in wb.sheetnames:
        title = wb[name].cell(row=1, column=1).value
        if title and INDICATOR_SHEET_TITLE_MARKER in str(title).strip().lower():
            return True
    return False


def _is_admin_activities_workbook(path: Path) -> bool:
    """Detected by sheet name ("Admin Activities") plus the task-column
    header text in A1 -- same "content, never filename" principle as every
    other domain, and the header check guards against an unrelated workbook
    that happens to reuse the sheet name."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return False
    if ADMIN_ACTIVITIES_SHEET not in wb.sheetnames:
        return False
    header = wb[ADMIN_ACTIVITIES_SHEET].cell(row=1, column=1).value
    return bool(header) and str(header).strip() == ADMIN_ACTIVITIES_HEADER


def detect_workbook_type(path: Path) -> DetectionResult:
    """Classify one .xlsx by its actual sheet names (Coverage/VPD) or, for
    the differently-shaped Indicator Sheet/Admin Activities workbooks, a
    content marker (A1 title / sheet name + header)."""
    try:
        sheets = _sheet_names(path)
    except Exception as e:
        return DetectionResult(
            "unreadable", [], [],
            f"{path.name!r} could not be opened as an Excel file ({e}). "
            f"Confirm it's a real, uncorrupted .xlsx workbook."
        )

    if not sheets:
        return DetectionResult("empty", [], [], f"{path.name!r} has no sheets -- it appears to be empty.")

    normalized = {s.strip().lower(): s for s in sheets}
    coverage_matches = [normalized[k] for k in normalized if k in COVERAGE_SIGNATURE]
    vpd_matches = [normalized[k] for k in normalized if k in VPD_SIGNATURE]
    who_matches = [normalized[k] for k in normalized if k in WHO_ACTIVITIES_SIGNATURE]

    if len(who_matches) >= len(WHO_ACTIVITIES_SIGNATURE):
        return DetectionResult(
            "who_activities", who_matches, sheets,
            f"Recognized as a WHO Supported Activities workbook (both {', '.join(sorted(WHO_ACTIVITIES_SIGNATURE))!r}-style sheets found)."
        )

    if not coverage_matches and not vpd_matches and _is_admin_activities_workbook(path):
        return DetectionResult(
            "admin_activities", [], sheets,
            f"Recognized as an Admin Activities checklist workbook (sheet name and header marker found)."
        )

    if not coverage_matches and not vpd_matches and _is_indicator_sheet_workbook(path):
        return DetectionResult(
            "indicator_sheet", [], sheets,
            f"Recognized as a Measles Indicator Sheet workbook (cell A1 title marker found)."
        )

    if len(coverage_matches) >= MIN_MATCHING_SHEETS and len(coverage_matches) >= len(vpd_matches):
        return DetectionResult(
            "coverage", coverage_matches, sheets,
            f"Recognized as a Coverage workbook ({len(coverage_matches)} of "
            f"{len(COVERAGE_SIGNATURE)} expected sheets found)."
        )
    if len(vpd_matches) >= MIN_MATCHING_SHEETS:
        return DetectionResult(
            "vpd", vpd_matches, sheets,
            f"Recognized as a VPD surveillance line list ({len(vpd_matches)} of "
            f"{len(VPD_SIGNATURE)} expected sheets found)."
        )
    shown = ", ".join(sheets[:6]) + (", ..." if len(sheets) > 6 else "")
    return DetectionResult(
        "unknown", [], sheets,
        f"Could not confidently identify {path.name!r} as a Coverage or VPD surveillance "
        f"workbook from its sheet names ({shown}). If this is meant to be a Coverage or VPD "
        f"file with renamed sheets, rename them to match the expected sheet names and "
        f"re-upload. Monitoring/supervisory-visit files (RCA / Supervisory Checklist) are "
        f"typically saved as .xls, not .xlsx -- see detect_monitoring_file."
    )


# Monitoring domain files (RCA / Supervisory Checklist) are HTML tables saved
# with a ".xls" extension, not real Excel workbooks -- openpyxl can't open
# them at all, so they need their own column-based signature, checked against
# an HTML-table read rather than sheet names.
RCA_COLUMN_SIGNATURE = {"record id", "child name", "penta 1", "vaccince source"}
SUPERVISORY_COLUMN_SIGNATURE = {
    "type of vaccintion site", "service functionality", "monitoring system quality",
}
MIN_MATCHING_COLUMNS = 2


def detect_monitoring_file(path: Path) -> DetectionResult:
    """Classify one Monitoring-domain ".xls" (HTML table) file by its actual
    column names. Never called on real .xlsx workbooks -- those go through
    detect_workbook_type instead."""
    try:
        tables = pd.read_html(path)
    except Exception as e:
        return DetectionResult(
            "unreadable", [], [],
            f"{path.name!r} could not be read as an HTML table ({e})."
        )
    if not tables or tables[0].empty:
        return DetectionResult("empty", [], [], f"{path.name!r} has no rows -- it appears to be empty.")

    columns = [str(c).strip() for c in tables[0].columns]
    normalized = {c.lower(): c for c in columns}
    rca_matches = [normalized[k] for k in normalized if k in RCA_COLUMN_SIGNATURE]
    supervisory_matches = [normalized[k] for k in normalized if k in SUPERVISORY_COLUMN_SIGNATURE]

    if len(rca_matches) >= MIN_MATCHING_COLUMNS and len(rca_matches) >= len(supervisory_matches):
        return DetectionResult(
            "rca", rca_matches, columns,
            f"Recognized as an RCA (Rapid Convenience Assessment) file ({len(rca_matches)} "
            f"of {len(RCA_COLUMN_SIGNATURE)} expected columns found)."
        )
    if len(supervisory_matches) >= MIN_MATCHING_COLUMNS:
        return DetectionResult(
            "supervisory", supervisory_matches, columns,
            f"Recognized as a Supervisory Checklist file ({len(supervisory_matches)} of "
            f"{len(SUPERVISORY_COLUMN_SIGNATURE)} expected columns found)."
        )
    shown = ", ".join(columns[:6]) + (", ..." if len(columns) > 6 else "")
    return DetectionResult(
        "unknown", [], columns,
        f"Could not confidently identify {path.name!r} as an RCA or Supervisory Checklist "
        f"file from its columns ({shown})."
    )
