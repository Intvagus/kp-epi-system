"""Read the "Divisional Officer & Admin Compliance" workbook (sheet name
"Admin Activities") -- a per-officer administrative compliance checklist
(19 fixed administrative responsibilities x 6 named officer columns, e.g.
logbook/report submission, claims processing, procurement, record
management), not a case-level or activity-log dataset like every other
domain in this project.

Each task belongs to one of two compliance sections -- "Divisional
Officer" or "Admin Section" -- or both (SECTION_BOTH); one task
("Other Assigned Tasks") belongs to both, tracked as two independent
checklist entries on the dashboard (see template.html). The user shared a
second reference workbook with a color-coding legend embedded in it (a
fill-color swatch in column J paired with a text label in column K, on a
few example rows) to specify this split: Yellow -> "Divisional Officer",
green (theme accent3) -> "Admin Section", magenta -> "Both", no fill ->
"Delete". Applying that legend to each task row's own column-A fill color
(never guessed at) gave the section for every task, and identified exactly
one task ("Monthly Report Compliance") as no-fill/"Delete" -- removed from
this workbook entirely, not just hidden. That same reference file also
renamed two travel-related tasks ("Duty Travel Requests" / "Travel Claims
/ Approval" -> "Travel Claims Submission" / "Travel Claims Processed"),
adopted here. The real workbook (data/raw/Admin_Activities_Checklist.xlsx)
was rebuilt to match: 19 tasks, a "Section" column added, the 6 real
officer names already in place kept as-is.

Confirmed by direct inspection: the source file received was originally a
BLANK CHECKLIST TEMPLATE with 4 generic "Officer 1".."Officer 4" columns,
not completed records -- every officer cell contained the literal
instructional string "Yes/No/NA" (telling whoever fills the sheet what to
type), not an actual answer, and the "Remarks / Evidence" column holds a
fixed descriptor of what evidence each task expects (e.g. "Monthly",
"Quality & completeness"), not a real per-period remark. There are no
dates, districts, UCs, activity types, or completed status values anywhere
in the file -- confirmed by checking every cell's value, fill color, and
comments (none carry hidden data). The user later provided the 6 real
officer names to use in place of the generic "Officer 1".."Officer 4"
labels; the source workbook's header row was updated in place to those 6
names (2 new officer columns added, same "Yes/No/NA" placeholder cells as
every other column) -- see CLAUDE.md for the exact names. All 120
(20 task x 6 officer) cells were still placeholder text at that point.

The user confirmed this genuinely is the intended source file and chose to
treat the Admin Activities tab as a LIVE, FILLABLE CHECKLIST inside the
dashboard (officers'/reviewers' Yes/No/N-A selections are entered and saved
in the browser, per the same localStorage-backed editable-note pattern used
elsewhere on this dashboard) rather than a chart-driven analytics tab, since
there is no completed data to chart. This module only reads the checklist's
real STRUCTURE from the workbook (task names, the officer column labels --
whatever names and however many are in the header row -- and each task's
expected-evidence descriptor) -- never the placeholder "Yes/No/NA" text as
if it were a real answer.

Forward-compatible: if a future version of this file DOES contain real
per-officer answers (a literal "Yes", "No", or "N/A", not the placeholder
phrase), _normalize_officer_cell recognizes and keeps them as genuine
source-provided starting values rather than discarding them -- see its own
docstring.
"""
import re
from pathlib import Path

import openpyxl

ADMIN_ACTIVITIES_SHEET = "Admin Activities"
TASK_COLUMN_HEADER = "Task / Administrative Responsibility"
EVIDENCE_COLUMN_HEADER = "Remarks / Evidence"
SECTION_COLUMN_HEADER = "Section"

# The two compliance sections this checklist is split into, plus the value
# a task carries when it belongs to both -- these are the exact strings the
# source workbook's own Section column uses (decoded from the color-coding
# legend embedded in the reference file the user shared: Yellow=Divisional
# Officer, green=Admin Section, magenta=Both, no-fill=Delete -- the "Delete"
# rows are simply not present in the workbook at all, never filtered here).
SECTION_DIVISIONAL_OFFICER = "Divisional Officer"
SECTION_ADMIN = "Admin Section"
SECTION_BOTH = "Both"

# The placeholder text a blank template cell contains, normalized (spaces/
# slashes stripped, lowercased) so "Yes/No/NA", "Yes / No / NA", etc. all
# match the same check.
_PLACEHOLDER_NORMALIZED = "yes/no/na".replace("/", "").replace(" ", "")


def _normalize_officer_cell(value):
    """None/blank -> None (unanswered). The literal placeholder phrase
    ("Yes/No/NA", however spaced) -> None (still unanswered -- it's
    instructional text, not a real answer). A genuine single-word answer
    ("Yes", "No", "NA"/"N/A", or abbreviations) -> the canonical "Yes" /
    "No" / "N/A". Anything else unrecognized is kept as-is (flagged by the
    caller via `officers_raw`), never silently dropped."""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    compact = re.sub(r"[\s/]+", "", s).lower()
    if compact == _PLACEHOLDER_NORMALIZED:
        return None
    normalized = s.strip().lower()
    if normalized in {"yes", "y"}:
        return "Yes"
    if normalized in {"no", "n"}:
        return "No"
    if normalized in {"na", "n/a", "not applicable"}:
        return "N/A"
    return s


def find_admin_activities_files(raw_dir: Path) -> list[Path]:
    files = []
    for path in sorted(raw_dir.glob("*.xlsx")):
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            continue
        if ADMIN_ACTIVITIES_SHEET in wb.sheetnames:
            header = wb[ADMIN_ACTIVITIES_SHEET].cell(row=1, column=1).value
            if header and str(header).strip() == TASK_COLUMN_HEADER:
                files.append(path)
    return files


def load_admin_activities(path: Path) -> dict:
    """Returns the checklist's real structure -- task list, officer column
    labels, per-task expected-evidence descriptor, and any genuine
    source-provided answers (see _normalize_officer_cell). Column layout is
    read from the header row itself (never hardcoded to exactly 4 officers),
    so a future version of this file with more/fewer officer columns is
    handled without a code change."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[ADMIN_ACTIVITIES_SHEET]

    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    header = [str(h).strip() if h is not None else None for h in header]
    task_col = header.index(TASK_COLUMN_HEADER) + 1 if TASK_COLUMN_HEADER in header else 1
    evidence_col = header.index(EVIDENCE_COLUMN_HEADER) + 1 if EVIDENCE_COLUMN_HEADER in header else None
    section_col = header.index(SECTION_COLUMN_HEADER) + 1 if SECTION_COLUMN_HEADER in header else None
    officer_cols = [
        (c, header[c - 1]) for c in range(1, ws.max_column + 1)
        if c not in (task_col, evidence_col, section_col) and header[c - 1]
    ]

    tasks = []
    prefilled_count = 0
    unrecognized_count = 0
    r = 2
    while True:
        task_name = ws.cell(row=r, column=task_col).value
        if task_name is None or str(task_name).strip() == "":
            break
        officers = {}
        for col, label in officer_cols:
            raw = ws.cell(row=r, column=col).value
            value = _normalize_officer_cell(raw)
            officers[label] = value
            if value is not None:
                if value in {"Yes", "No", "N/A"}:
                    prefilled_count += 1
                else:
                    unrecognized_count += 1
        tasks.append({
            "task": str(task_name).strip(),
            "expected_evidence": (str(ws.cell(row=r, column=evidence_col).value).strip()
                                   if evidence_col and ws.cell(row=r, column=evidence_col).value else None),
            "section": (str(ws.cell(row=r, column=section_col).value).strip()
                        if section_col and ws.cell(row=r, column=section_col).value else None),
            "officers": officers,
        })
        r += 1

    return {
        "source_file": path.name,
        "officer_labels": [label for _, label in officer_cols],
        "tasks": tasks,
        "prefilled_answers_count": prefilled_count,
        "unrecognized_cell_count": unrecognized_count,
    }


def build_summary(loaded: dict) -> dict:
    return {
        "status": "ok",
        "source_file": loaded["source_file"],
        "officer_labels": loaded["officer_labels"],
        "tasks": loaded["tasks"],
        "task_count": len(loaded["tasks"]),
        "officer_count": len(loaded["officer_labels"]),
        "prefilled_answers_count": loaded["prefilled_answers_count"],
        "unrecognized_cell_count": loaded["unrecognized_cell_count"],
        "is_blank_template": loaded["prefilled_answers_count"] == 0,
    }


PROJECT_ROOT = Path(__file__).resolve().parents[2]
PROCESSED_DIR = PROJECT_ROOT / "data" / "processed"


def run_admin_activities(raw_dir: Path | None = None, processed_dir: Path | None = None) -> dict | None:
    """Independent of every other domain -- see every other run_* entry
    point in this project for the same convention."""
    import json

    raw_dir = raw_dir or (PROJECT_ROOT / "data" / "raw")
    processed_dir = processed_dir or PROCESSED_DIR
    files = find_admin_activities_files(raw_dir)
    if not files:
        return None

    print("\nAdmin Activities pipeline starting...")
    path = files[0]
    print(f"  Loading {path.name}...")
    loaded = load_admin_activities(path)
    summary = build_summary(loaded)
    print(f"  {summary['task_count']} tasks x {summary['officer_count']} officers "
          f"({summary['prefilled_answers_count']} pre-filled answers found in source)")

    processed_dir.mkdir(parents=True, exist_ok=True)
    with open(processed_dir / "admin_activities_summary.json", "w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2, default=str)
    print("Admin Activities pipeline finished OK.")
    return summary
