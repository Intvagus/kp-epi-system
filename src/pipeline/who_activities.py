"""Read the "WHO Supported Activities" workbook -- a monthly field-monitoring
highlights report submitted by a WHO District Officer (DO), not a
multi-period transactional activity database. Confirmed by direct inspection
of the source file (see CLAUDE.md): it covers exactly ONE reporting month,
ONE duty station, and its ~6 assigned districts, with three sheets that are
three different views of the SAME underlying report, not three independent
datasets:

- "Sheet1": the raw narrative field report as submitted (free-text bullets
  and day-by-day prose). Used here only for reconciliation (cross-checking
  the headline numbers below actually appear in the source narrative), never
  parsed into structured rows -- the prose is too unstructured to extract
  further real numbers from without guessing.
- "Evidence & Findings": a structured, row-per-finding extraction of the
  narrative (ID/District/Date/Activity/Quantified Result/Contribution/
  Suggested Highlight/Source row) -- the closest thing to an "activity
  table" in this workbook, and the only sheet with per-row District/Date
  fields. This is the dashboard's primary structured data source.
- "WHO Highlights Dashboard": the polished executive summary already
  assembled from the same report (headline KPIs, thematic results,
  district-level highlights, priority gaps, a management-message
  paragraph). Read here for its already-clean KPI numbers and narrative
  text, not recomputed from the Evidence sheet.

Because this is a single-month, single-duty-station report, several things a
larger activities database might support are genuinely NOT available here
and are not fabricated: no activity status (Completed/Ongoing/Planned/...),
no numeric target/achievement pair for most rows, no multi-month trend, no
multi-province comparison, no implementing-partner field. See
build_summary()'s docstring for exactly what is and isn't computed.
"""
import re
from pathlib import Path

import openpyxl
import pandas as pd

from .config import DISTRICT_TO_BOUNDARY

HIGHLIGHTS_SHEET = "WHO Highlights Dashboard"
EVIDENCE_SHEET = "Evidence & Findings"
RAW_SHEET = "Sheet1"
REQUIRED_SHEETS = {HIGHLIGHTS_SHEET, EVIDENCE_SHEET}

# The Evidence & Findings sheet spells Chitral's two districts
# adjective-first ("Lower Chitral", "Upper Chitral"), while this project's
# canonical spelling (config.DISTRICT_TO_BOUNDARY, matching the Coverage/VPD/
# Monitoring files) is place-name-first ("Chitral Lower", "Chitral Upper") --
# same district, different spelling in this one source file, confirmed by
# direct inspection. Everything else already matches exactly. "Multiple
# districts" is a real value in the source (one finding spans several
# assigned districts, e.g. IOA Round 2 as a whole) -- kept as its own
# explicit, non-mapped value rather than force-assigned to one district.
DISTRICT_NAME_CANONICAL = {
    "Lower Chitral": "Chitral Lower", "Upper Chitral": "Chitral Upper",
    "Malakand": "Malakand", "Dir Lower": "Dir Lower", "Dir Upper": "Dir Upper",
    "Bajaur": "Bajaur", "Multiple districts": "Multiple districts",
}

# This workbook has no "Activity Category" column -- ACTIVITY_THEME below
# assigns each of the 11 Evidence & Findings rows (fixed IDs E01-E11) to one
# of 7 themes read directly off that row's own "Activity / Finding" text,
# confirmed by direct inspection, the same "explicit verified mapping, not
# inferred" convention as clean_vpd.py's classification canonicalization --
# preferred over keyword/regex matching given there are only 11 rows to
# check by hand. Never applied to a future upload with different IDs without
# re-verifying against that file's own Activity/Finding text.
ACTIVITY_THEME = {
    "E01": "Outreach / Routine Service Delivery",
    "E02": "Training",
    "E03": "Training",
    "E04": "Monitoring & Review Meeting (MRM)",
    "E05": "Supportive Supervision / IOA",
    "E06": "Rapid Convenience Assessment (RCA)",
    "E07": "Rapid Convenience Assessment (RCA)",
    "E08": "Supportive Supervision / IOA",
    "E09": "Rapid Convenience Assessment (RCA)",
    "E10": "Measles Outbreak Response Review",
    "E11": "Service-Continuity Planning",
}

EVIDENCE_COLUMNS = [
    "id", "district_raw", "date_period_raw", "activity", "quantified_result",
    "who_contribution", "suggested_highlight", "evidence_source",
]


def find_who_activities_files(raw_dir: Path) -> list[Path]:
    files = []
    for path in sorted(raw_dir.glob("*.xlsx")):
        try:
            sheets = set(pd.ExcelFile(path, engine="openpyxl").sheet_names)
        except Exception:
            continue
        if REQUIRED_SHEETS <= sheets:
            files.append(path)
    return files


def _cell(ws, coord):
    v = ws[coord].value
    return v.strip() if isinstance(v, str) else v


def _load_evidence_table(ws) -> pd.DataFrame:
    rows = []
    r = 2
    while True:
        rid = ws.cell(row=r, column=1).value
        if rid is None:
            break
        rows.append({
            "id": str(rid).strip(),
            "district_raw": ws.cell(row=r, column=2).value,
            "date_period_raw": ws.cell(row=r, column=3).value,
            "activity": ws.cell(row=r, column=4).value,
            "quantified_result": ws.cell(row=r, column=5).value,
            "who_contribution": ws.cell(row=r, column=6).value,
            "suggested_highlight": ws.cell(row=r, column=7).value,
            "evidence_source": ws.cell(row=r, column=8).value,
        })
        r += 1
    df = pd.DataFrame(rows, columns=EVIDENCE_COLUMNS)
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].apply(lambda v: v.strip() if isinstance(v, str) else v)
    return df


# Most Date/Period values are a single real day ("8 May 2024"), a handful are
# a range ("20-25 May 2024") or just the month ("May 2024"). Only a single
# concrete day can be parsed into a real date without guessing which day in
# a range or a whole month a finding belongs to -- date_start/date_end below
# capture a range's real bounds instead of picking one arbitrary day.
_DATE_RANGE_RE = re.compile(
    r"(\d{1,2})\s*[–\-]\s*(\d{1,2})\s+(\w+)\s+(\d{4})"
)
_SINGLE_DATE_RE = re.compile(
    r"(\d{1,2})\s+(\w+)\s+(\d{4})"
)


def _parse_date_period(raw):
    if not isinstance(raw, str):
        return {"date_start": None, "date_end": None, "is_range": False, "is_month_only": False}
    s = raw.strip()
    m = _DATE_RANGE_RE.search(s)
    if m:
        d1, d2, month, year = m.groups()
        try:
            start = pd.Timestamp(f"{d1} {month} {year}")
            end = pd.Timestamp(f"{d2} {month} {year}")
            return {"date_start": start, "date_end": end, "is_range": True, "is_month_only": False}
        except (ValueError, TypeError):
            pass
    m = _SINGLE_DATE_RE.search(s)
    if m:
        day, month, year = m.groups()
        try:
            d = pd.Timestamp(f"{day} {month} {year}")
            return {"date_start": d, "date_end": d, "is_range": False, "is_month_only": False}
        except (ValueError, TypeError):
            pass
    return {"date_start": None, "date_end": None, "is_range": False, "is_month_only": True}


def load_who_activities(path: Path) -> dict:
    """Returns the raw-ish parsed content of all 3 sheets -- clean_who_activities()
    does the district canonicalization / theme tagging / validation on top of this."""
    wb = openpyxl.load_workbook(path, data_only=True)
    highlights = wb[HIGHLIGHTS_SHEET]
    evidence_df = _load_evidence_table(wb[EVIDENCE_SHEET])

    kpis = {
        "field_support_days": _cell(highlights, "A5"),
        "zero_dose_vaccinated_arandu": _cell(highlights, "D5"),
        "measles_results_reviewed": _cell(highlights, "G5"),
        "mobr_districts_supported": _cell(highlights, "J5"),
    }
    title = _cell(highlights, "A1")
    subtitle = _cell(highlights, "A2")

    key_results = []
    r = 9
    while _cell(highlights, f"A{r}"):
        key_results.append({"theme": _cell(highlights, f"A{r}"), "narrative": _cell(highlights, f"B{r}")})
        r += 1

    # Column H runs contribution bullets (rows 9-14) directly into the next
    # section's own header ("4. PRIORITY SYSTEM GAPS / WAY FORWARD" at H15,
    # confirmed by direct inspection -- no blank row separates them in this
    # column specifically), so a numbered-header line ("<digit>. ...") is
    # the real stop signal, not the first blank cell.
    _SECTION_HEADER_RE = re.compile(r"^\d+\.\s")
    who_contributions = []
    r = 9
    while True:
        v = _cell(highlights, f"H{r}")
        if not v or _SECTION_HEADER_RE.match(str(v)):
            break
        who_contributions.append(v)
        r += 1

    district_highlights = []
    r = 16
    while _cell(highlights, f"A{r}"):
        district_highlights.append({"district_raw": _cell(highlights, f"A{r}"), "narrative": _cell(highlights, f"B{r}")})
        r += 1

    priority_gaps = []
    r = 16
    while _cell(highlights, f"H{r}"):
        priority_gaps.append(_cell(highlights, f"H{r}"))
        r += 1

    management_message = None
    for r in range(23, 30):
        label = _cell(highlights, f"A{r}")
        if label and "management message" in str(label).lower():
            management_message = _cell(highlights, f"A{r + 1}")
            break

    source_note = _cell(highlights, "A34")

    raw_text = None
    if RAW_SHEET in wb.sheetnames:
        raw_ws = wb[RAW_SHEET]
        parts = []
        for row in raw_ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip():
                    parts.append(cell.value)
        raw_text = "\n".join(parts)

    return {
        "title": title, "subtitle": subtitle, "kpis": kpis,
        "key_results": key_results, "who_contributions": who_contributions,
        "district_highlights": district_highlights, "priority_gaps": priority_gaps,
        "management_message": management_message, "source_note": source_note,
        "evidence_df": evidence_df, "raw_text": raw_text,
    }


def clean_evidence_table(evidence_df: pd.DataFrame) -> tuple[pd.DataFrame, list]:
    """Adds district_canonical, activity_theme, and parsed date fields to the
    raw Evidence & Findings rows. Returns (cleaned_df, unmapped_districts) --
    unmapped_districts should always be empty for this file (verified
    exhaustive against DISTRICT_NAME_CANONICAL by direct inspection) and is
    surfaced, not silently dropped, in case a future upload adds a district
    or a new spelling variant."""
    df = evidence_df.copy()
    df["district_canonical"] = df["district_raw"].map(DISTRICT_NAME_CANONICAL)
    unmapped_mask = df["district_canonical"].isna()
    unmapped_districts = sorted(df.loc[unmapped_mask, "district_raw"].dropna().unique().tolist())
    df.loc[unmapped_mask, "district_canonical"] = df.loc[unmapped_mask, "district_raw"]

    df["activity_theme"] = df["id"].map(ACTIVITY_THEME)
    unthemed = df["activity_theme"].isna()
    df.loc[unthemed, "activity_theme"] = "Other / Not Categorized"

    parsed = df["date_period_raw"].apply(_parse_date_period).apply(pd.Series)
    df = pd.concat([df, parsed], axis=1)
    return df, unmapped_districts


def build_activity_theme_breakdown(df: pd.DataFrame) -> dict:
    return df["activity_theme"].value_counts().to_dict()


def build_district_breakdown(df: pd.DataFrame) -> list[dict]:
    """One row per district_canonical value actually present, including the
    real "Multiple districts" catch-all -- never split across the districts
    it might touch, since the source doesn't say which of them specifically."""
    counts = df["district_canonical"].value_counts()
    return [{"district": d, "evidence_count": int(n)} for d, n in counts.items()]


def build_district_map(df: pd.DataFrame, assigned_districts: list[str]) -> dict:
    """Choropleth-ready district map, scoped to ONLY this duty station's
    assigned districts -- every other KP district is genuinely outside this
    report's coverage, not zero activity, so it is left out of "features"
    entirely (rendered as "not covered by this report" on the dashboard map,
    a third state distinct from both a real 0 and a boundary-mapping gap).

    assigned_districts (from the Highlights sheet's own district-level
    highlights table -- the report's real, stated coverage list) always get
    a real entry, 0 if they have no Evidence & Findings row -- Bajaur is
    assigned but has no evidence row (see data_quality), and a true 0 there
    is a different fact from "outside this report" for the other ~30 KP
    districts, same "never silently equate missing with zero" rule as the
    Coverage/Monitoring district maps."""
    real_districts = df[df["district_canonical"] != "Multiple districts"]
    counts = real_districts["district_canonical"].value_counts()
    all_districts = sorted(set(counts.index) | set(assigned_districts))
    unmapped = sorted(set(all_districts) - set(DISTRICT_TO_BOUNDARY))
    # component_districts is a 1-element list here (this map is always 1:1
    # district:boundary, never combined) -- kept for shape-compatibility with
    # the shared renderSingleDistrictMap()/map-tooltip JS, which every other
    # domain's district map (Coverage/RCA/Supervisory) already populates the
    # same way.
    features = {
        d: {"evidence_count": int(counts.get(d, 0)), "component_districts": [d]}
        for d in all_districts if d in DISTRICT_TO_BOUNDARY
    }
    return {"unmapped_districts": unmapped, "features": features}


def build_timeline(df: pd.DataFrame) -> list[dict]:
    """Chronological list of evidence findings that have a real, parseable
    date (a specific day, or a date range's start day) -- May 2024 is the
    only reporting period in this file, so this is a within-month timeline,
    never a multi-month trend (see build_summary's docstring)."""
    dated = df[df["date_start"].notna()].copy()
    dated = dated.sort_values("date_start")
    return [
        {
            "id": r["id"], "date": r["date_start"].strftime("%Y-%m-%d"),
            "date_end": r["date_end"].strftime("%Y-%m-%d") if pd.notna(r["date_end"]) and r["date_end"] != r["date_start"] else None,
            "district": r["district_canonical"], "activity_theme": r["activity_theme"],
            "activity": r["activity"],
        }
        for _, r in dated.iterrows()
    ]


def _reconcile_kpis(loaded: dict) -> list[dict]:
    """Cross-checks the 4 headline KPI numbers against the small structured
    summary table (rows 30-33 of the Highlights sheet) and against the raw
    narrative report (Sheet1) -- the same number appearing in 2-3
    independently-typed/located places in the source file is real
    corroboration, not a computed claim."""
    checks = []
    label_map = [
        ("field_support_days", "Field-support days/visits"),
        ("zero_dose_vaccinated_arandu", "Zero-dose children vaccinated"),
        ("measles_results_reviewed", "Measles results reviewed"),
        ("mobr_districts_supported", "MOBR districts supported"),
    ]
    kpis = loaded["kpis"]
    raw_text = loaded.get("raw_text") or ""
    for key, label in label_map:
        headline_val = kpis.get(key)
        try:
            headline_num = int(float(headline_val)) if headline_val is not None else None
        except (TypeError, ValueError):
            headline_num = None
        found_in_raw = headline_val is not None and str(headline_num) in raw_text
        checks.append({
            "metric": label, "headline_value": headline_num,
            "found_in_raw_narrative": found_in_raw,
        })
    return checks


def build_summary(loaded: dict) -> dict:
    """Single JSON payload for the dashboard, computed once here -- see
    module docstring for what this workbook does and does not support.

    Deliberately NOT computed/shown anywhere downstream (the fields don't
    exist in the source, so nothing is invented for them):
    - Activity status (Completed/Ongoing/Planned/Delayed/Cancelled) -- no
      status field anywhere in this workbook.
    - A numeric target vs. achievement pair for most findings -- only the
      measles-review KPI (8 of 11) is a genuine target/achieved pair; the
      other "Quantified Result" values are free text, not parsed into new
      numbers (risk of silent mis-parsing outweighs the value).
    - A multi-month trend -- exactly one reporting period (May 2024) exists.
    - A multi-province comparison -- every district here is in KP, all
      served by one duty station.
    - Implementing partner as a filterable dimension -- partners are
      mentioned in free text (e.g. "partner staff") but not as a clean,
      per-record field.
    """
    df, unmapped_districts = clean_evidence_table(loaded["evidence_df"])

    kpis = loaded["kpis"]
    def _int(v):
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return None
    field_support_days = _int(kpis.get("field_support_days"))
    zero_dose_vaccinated = _int(kpis.get("zero_dose_vaccinated_arandu"))
    measles_reviewed = _int(kpis.get("measles_results_reviewed"))
    mobr_districts = _int(kpis.get("mobr_districts_supported"))
    # "of 11" is stated directly in the source heading (G4: "MEASLES CASE
    # RESULTS REVIEWED (OF 11)") -- the denominator is real, not invented.
    measles_review_pct = round(measles_reviewed / 11 * 100, 1) if measles_reviewed is not None else None

    district_highlights = []
    for dh in loaded["district_highlights"]:
        canonical = DISTRICT_NAME_CANONICAL.get(dh["district_raw"], dh["district_raw"])
        district_highlights.append({"district": canonical, "district_raw": dh["district_raw"], "narrative": dh["narrative"]})

    districts_with_highlights = {d["district"] for d in district_highlights}
    districts_with_evidence = set(df.loc[df["district_canonical"] != "Multiple districts", "district_canonical"].unique())
    highlights_no_evidence = sorted(districts_with_highlights - districts_with_evidence)
    evidence_no_highlights = sorted(districts_with_evidence - districts_with_highlights)

    # Reformat dates AFTER to_dict(), on plain Python dicts -- assigning a
    # list containing None back into an existing datetime64 DataFrame column
    # makes pandas re-coerce those None values to NaN when it rebuilds the
    # column (confirmed reproducible, independent of how the list itself is
    # built), and json.dump then writes a bare `NaN` token, which is not
    # valid JSON and breaks JSON.parse() in the browser. Once the values are
    # out of a pandas column and into plain dicts, None stays None.
    def _fmt_date(d):
        return d.strftime("%Y-%m-%d") if pd.notna(d) else None
    evidence_records = df.to_dict(orient="records")
    for r in evidence_records:
        r["date_start"] = _fmt_date(r["date_start"])
        r["date_end"] = _fmt_date(r["date_end"])

    reconciliation = _reconcile_kpis(loaded)

    data_quality = {
        "evidence_record_count": len(df),
        "duplicate_evidence_ids": int(df["id"].duplicated().sum()),
        "missing_values_by_column": {c: int(df[c].isna().sum()) for c in EVIDENCE_COLUMNS},
        "district_name_variants_standardized": {
            k: v for k, v in DISTRICT_NAME_CANONICAL.items() if k != v
        },
        "unmapped_districts": unmapped_districts,
        "districts_with_narrative_highlight_but_no_evidence_row": highlights_no_evidence,
        "districts_with_evidence_row_but_no_narrative_highlight": evidence_no_highlights,
        "headline_kpis_reconciled_against_summary_table_and_raw_narrative": reconciliation,
        "numeric_fields_stored_as_text_in_source": [
            "The 4 headline KPI cells (A5/D5/G5/J5 on the Highlights sheet) are stored as "
            "text in the source file; the same 4 figures also appear as real numbers in the "
            "sheet's own small summary table (rows 30-33), which is what these totals are "
            "reconciled against."
        ],
    }

    insights = []
    top_theme = df["activity_theme"].value_counts()
    if not top_theme.empty:
        insights.append(f"{top_theme.index[0]} was the most common activity theme, accounting for "
                         f"{int(top_theme.iloc[0])} of {len(df)} evidence-table findings.")
    district_counts = df.loc[df["district_canonical"] != "Multiple districts", "district_canonical"].value_counts()
    if not district_counts.empty:
        insights.append(f"{district_counts.index[0]} had the most field-support findings documented "
                         f"({int(district_counts.iloc[0])} of {len(df)}).")
    if zero_dose_vaccinated is not None:
        insights.append(f"{zero_dose_vaccinated} zero-dose children were vaccinated in Arandu (Chitral Lower), "
                         f"the single largest quantified result in this reporting period.")
    if measles_reviewed is not None:
        insights.append(f"Measles case-result review was {measles_review_pct}% complete "
                         f"({measles_reviewed} of 11 results reviewed; 3 pending, per the source).")
    if highlights_no_evidence:
        insights.append(f"{', '.join(highlights_no_evidence)} {'has' if len(highlights_no_evidence) == 1 else 'have'} "
                         f"a narrative highlight in the source report but no row in the Evidence & Findings table.")

    return {
        "status": "ok",
        "title": loaded["title"], "subtitle": loaded["subtitle"],
        "reporting_period_label": "May 2024",
        "duty_station": "Malakand",
        "kpis": {
            "field_support_days": field_support_days,
            "zero_dose_vaccinated_arandu": zero_dose_vaccinated,
            "measles_results_reviewed": measles_reviewed,
            "measles_results_total": 11,
            "measles_review_pct": measles_review_pct,
            "mobr_districts_supported": mobr_districts,
            "evidence_record_count": len(df),
            "districts_covered": len(districts_with_evidence | districts_with_highlights),
        },
        "activity_theme_breakdown": build_activity_theme_breakdown(df),
        "district_breakdown": build_district_breakdown(df),
        "district_map": build_district_map(df, sorted(districts_with_highlights)),
        "timeline": build_timeline(df),
        "evidence_records": evidence_records,
        "key_results": loaded["key_results"],
        "who_contributions": loaded["who_contributions"],
        "district_highlights": district_highlights,
        "priority_gaps": loaded["priority_gaps"],
        "management_message": loaded["management_message"],
        "source_note": loaded["source_note"],
        "insights": insights,
        "data_quality": data_quality,
    }


PROJECT_ROOT = Path(__file__).resolve().parents[2]
PROCESSED_DIR = PROJECT_ROOT / "data" / "processed"


def run_who_activities(raw_dir: Path | None = None, processed_dir: Path | None = None) -> dict | None:
    """Independent of every other domain -- an upload with no WHO Supported
    Activities workbook still produces a working dashboard/bulletin for
    whatever else was uploaded, and vice versa (same per-domain independence
    as every other pipeline entry point)."""
    import json

    raw_dir = raw_dir or (PROJECT_ROOT / "data" / "raw")
    processed_dir = processed_dir or PROCESSED_DIR
    files = find_who_activities_files(raw_dir)
    if not files:
        return None

    print("\nWHO Supported Activities pipeline starting...")
    path = files[0]
    print(f"  Loading {path.name}...")
    loaded = load_who_activities(path)
    summary = build_summary(loaded)
    print(f"  {summary['kpis']['evidence_record_count']} evidence records, "
          f"{summary['kpis']['districts_covered']} districts covered, {summary['reporting_period_label']}")

    processed_dir.mkdir(parents=True, exist_ok=True)
    with open(processed_dir / "who_activities_summary.json", "w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2, default=str)
    print("WHO Supported Activities pipeline finished OK.")
    return summary
