"""Read the "Measles Indicator Sheet" workbook -- a separate, real .xlsx
export (not the HTML-table Monitoring files) containing one sheet per year,
each a per-district table of measles/rubella surveillance performance
indicators plus a "Provincial Total" row. Distinct from the MSL line list
(KP VPDs Line List Week N-N,YYYY.xlsx): the line list is case-level raw
data, this sheet is the source system's own pre-aggregated indicator
report, with a different (much smaller) district total than the line list's
weekly-cumulative count -- the two are never reconciled against each other,
only shown side by side.

Column layout (row 2 = header, row 3 = sub-header for the G:K "Measles
Cases" group, data from row 4, "Provincial Total" as the last data row) is
confirmed via direct inspection of the source workbook's cell fills: columns
F, L, N, P, Q, S are highlighted a distinct color in the source (rate/%
indicators) vs. B/C/D/E/G-K/M/O/R (raw counts) -- those 6 highlighted
columns are the "key indicators" this module surfaces, never invented.
"""
import json
from pathlib import Path

import openpyxl

INDICATOR_SHEET_TITLE_MARKER = "indicator sheet"

# This workbook spells district names differently from the Coverage/
# Monitoring files (its own data-entry convention, confirmed by direct
# inspection -- e.g. "Bajour" not "Bajaur", "D I Khan" not "D.I. Khan").
# Maps every one of its 37 district rows (36 real + "Torghar", matching
# DISTRICT_TO_BOUNDARY's 36 keys plus the real Tor Ghar district that has no
# Coverage-file data) to this project's canonical spelling, so the incidence
# map can share the same kp_districts.geojson boundaries as every other map.
# "South Wazirisan Upper"/"South Waziristan Lower" -> SW Wazir Belt/SW
# Mehsud Belt carries the same Upper=Wazir/Lower=Mehsud assumption flagged in
# config.DISTRICT_TO_BOUNDARY -- this sheet's own Upper/Lower naming is a
# second, independent data point consistent with that assumption, not proof.
DISTRICT_NAME_CANONICAL = {
    "Abbottabad": "Abbottabad", "Bajour": "Bajaur", "Bannu": "Bannu",
    "Battagram": "Battagram", "Buner": "Buner", "Charssada": "Charsadda",
    "Chitral Lower": "Chitral Lower", "Chitral Upper": "Chitral Upper",
    "D I Khan": "D.I. Khan", "Dir Lower": "Dir Lower", "Dir Upper": "Dir Upper",
    "Hangu": "Hangu", "Haripur": "Haripur", "Karak": "Karak", "Khyber": "Khyber",
    "Kohat": "Kohat", "Kohistan Lower": "Kohistan Lower", "Kohistan Upper": "Kohistan Upper",
    "Kolai Palas": "Kolai Palas Kohistan", "Kurram L&C": "Kurram Lower and Central",
    "Kurram Upper": "Kurram Upper", "Lakki Marwat": "Lakki Marwat", "Malakand": "Malakand",
    "Mansehra": "Mansehra", "Mardan": "Mardan", "Mohmand": "Mohmand",
    "North Waziristan": "North Waziristan", "Nowshera": "Nowshera", "Orakzai": "Orakzai",
    "Peshawar": "Peshawar", "Shangla": "Shangla",
    "South Wazirisan Upper": "SW Wazir Belt", "South Waziristan Lower": "SW Mehsud Belt",
    "Swabi": "Swabi", "Swat": "Swat", "Tank": "Tank", "Torghar": "Tor Ghar",
}

# WHO-standard measles incidence bands (cases per million population,
# annualized -- this sheet's own "measles_incidence_per_million" column is
# already an annualized rate, confirmed by its column header). Fixed
# thresholds, not derived from the data.
INCIDENCE_BANDS = [
    (5, "green", "Low (<5)"),
    (20, "yellow", "Moderate (5 to <20)"),
    (None, "red", "Disruptive outbreak (≥20)"),
]


def _incidence_band(value):
    if value is None:
        return None
    for upper, color, label in INCIDENCE_BANDS:
        if upper is None or value < upper:
            return {"color": color, "label": label}
    return None

# Fixed column positions (1-indexed), confirmed against the real workbook --
# not name-based, since the header text has trailing whitespace / minor
# year-to-year wording drift ("% Sample Collected" vs "% Sampling").
COLUMNS = {
    "district": 1, "total_population": 2, "minimum_expected_cases": 3,
    "total_cases_reported": 4, "non_measles_non_rubella_cases": 5,
    "non_measles_non_rubella_rate": 6,
    "measles_lab_confirmed": 7, "measles_double_infection": 8,
    "measles_epi_linked": 9, "measles_clinically_compatible": 10, "measles_total": 11,
    "measles_incidence_per_million": 12, "rubella_confirmed_cases": 13,
    "rubella_incidence_per_million": 14, "pending_classification": 15,
    "pct_sample_collected": 16, "pct_adequate_investigation": 17,
    "total_deaths": 18, "measles_related_deaths": 19,
}
HEADER_ROW = 2
DATA_START_ROW = 4
PROVINCIAL_TOTAL_LABEL = "Provincial Total"

# The 6 columns confirmed highlighted in the source (a distinct fill color
# from every other column) -- the sheet author's own designation of which
# indicators matter most, not a selection made in this codebase.
KEY_INDICATORS = [
    {"key": "non_measles_non_rubella_rate", "label": "Non-Measles/Non-Rubella (Discard) Rate",
     "unit": "per 100,000 population (annualized)", "decimals": 1},
    {"key": "measles_incidence_per_million", "label": "Measles Incidence",
     "unit": "per million population (annualized)", "decimals": 1},
    {"key": "rubella_incidence_per_million", "label": "Rubella Incidence",
     "unit": "per million population", "decimals": 1},
    {"key": "pct_sample_collected", "label": "Sample Collection",
     "unit": "% of cases", "decimals": 1},
    {"key": "pct_adequate_investigation", "label": "Adequate Investigation",
     "unit": "% of cases", "decimals": 1},
    {"key": "measles_related_deaths", "label": "Measles-Related Deaths",
     "unit": "count", "decimals": 0},
]


def find_indicator_sheet_files(raw_dir: Path) -> list[Path]:
    files = []
    for path in sorted(raw_dir.glob("*.xlsx")):
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            continue
        for name in wb.sheetnames:
            title = wb[name].cell(row=1, column=1).value
            if title and INDICATOR_SHEET_TITLE_MARKER in str(title).strip().lower():
                files.append(path)
                break
    return files


def _select_year_sheet(wb) -> tuple[str, object]:
    """Pick the most recent year-named sheet whose header row actually
    matches the expected layout -- some older years in this workbook use a
    different table shape entirely, so structural validation (not just the
    sheet existing) decides which one is usable."""
    year_sheets = sorted((n for n in wb.sheetnames if n.strip().isdigit()), reverse=True)
    for name in year_sheets:
        ws = wb[name]
        if (str(ws.cell(row=HEADER_ROW, column=COLUMNS["district"]).value or "").strip().lower() == "district"
                and "population" in str(ws.cell(row=HEADER_ROW, column=COLUMNS["total_population"]).value or "").lower()):
            return name, ws
    raise ValueError(
        "No year sheet in the Indicator workbook has the expected 'District' / "
        "'Total Population' header layout -- check the sheet structure or update "
        "src/pipeline/indicator_sheet_vpd.py's COLUMNS mapping."
    )


def load_indicator_sheet(path: Path) -> dict:
    """Returns {'year': ..., 'districts': [...], 'provincial_total': {...}}
    -- every row read positionally per COLUMNS, no formula recomputation
    (the sheet's own values, including its own 'Provincial Total' row, are
    trusted as-is, same 'trust the sheet' rule as Coverage's UC-level
    Access/Utilisation)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    year, ws = _select_year_sheet(wb)

    def _row(r):
        return {key: ws.cell(row=r, column=col).value for key, col in COLUMNS.items()}

    districts = []
    provincial_total = None
    r = DATA_START_ROW
    while True:
        district_name = ws.cell(row=r, column=COLUMNS["district"]).value
        if district_name is None:
            break
        row = _row(r)
        if str(district_name).strip() == PROVINCIAL_TOTAL_LABEL:
            provincial_total = row
            break
        districts.append(row)
        r += 1

    if provincial_total is None:
        raise ValueError(
            f"No '{PROVINCIAL_TOTAL_LABEL}' row found below the district rows in the "
            f"'{year}' sheet -- the workbook layout may have changed."
        )
    return {"year": year, "districts": districts, "provincial_total": provincial_total}


def build_key_indicators_summary(sheet: dict) -> dict:
    """One row per highlighted key indicator: the Provincial Total's own
    value (trusted as-is), plus the highest/lowest-value district for that
    indicator among the real per-district rows -- both genuinely sourced
    from the workbook, not a fabricated 'target'. No numeric target exists
    anywhere in this workbook (confirmed by inspection), so none is
    invented here; the dashboard shows that fact explicitly rather than a
    guessed WHO benchmark."""
    districts = sheet["districts"]
    prov = sheet["provincial_total"]
    rows = []
    for spec in KEY_INDICATORS:
        key = spec["key"]
        valued = [(d["district"], d[key]) for d in districts if isinstance(d[key], (int, float))]
        highest = max(valued, key=lambda t: t[1]) if valued else None
        lowest = min(valued, key=lambda t: t[1]) if valued else None
        rows.append({
            "key": key, "label": spec["label"], "unit": spec["unit"],
            "provincial_value": prov[key],
            "highest_district": {"district": highest[0], "value": highest[1]} if highest else None,
            "lowest_district": {"district": lowest[0], "value": lowest[1]} if lowest else None,
        })
    return {
        "status": "ok",
        "year": sheet["year"],
        "districts_covered": len(districts),
        "total_population": prov["total_population"],
        "total_cases_reported": prov["total_cases_reported"],
        "indicators": rows,
        "measles_incidence_map": build_measles_incidence_map(sheet),
    }


def build_measles_incidence_map(sheet: dict) -> dict:
    """Per-district measles incidence (cases per million, this sheet's own
    annualized figure -- never derived from a population proxy elsewhere in
    this project, see CLAUDE.md's "Confirmed VPD decisions") banded into the
    WHO-standard Low/Moderate/Disruptive-outbreak categories, keyed by this
    project's canonical district name so it shares kp_districts.geojson with
    every other choropleth map."""
    features = {}
    unmapped = []
    for row in sheet["districts"]:
        raw_name = str(row["district"]).strip()
        canonical = DISTRICT_NAME_CANONICAL.get(raw_name)
        if canonical is None:
            unmapped.append(raw_name)
            continue
        value = row["measles_incidence_per_million"]
        value = value if isinstance(value, (int, float)) else None
        band = _incidence_band(value)
        features[canonical] = {
            "measles_incidence_per_million": value,
            "category": band["color"] if band else None,
            "category_label": band["label"] if band else None,
        }
    return {"unmapped_districts": sorted(unmapped), "features": features}


PROJECT_ROOT = Path(__file__).resolve().parents[2]
PROCESSED_DIR = PROJECT_ROOT / "data" / "processed"


def run_indicator_sheet(raw_dir: Path | None = None, processed_dir: Path | None = None) -> dict | None:
    """Independent of the MSL line list pipeline (run_vpd.py) -- an
    Indicator Sheet upload with no line list present still produces
    data/processed/vpd_indicator_summary.json, and vice versa. Returns None
    (writes nothing) if no Indicator Sheet workbook is found."""
    raw_dir = raw_dir or (PROJECT_ROOT / "data" / "raw")
    processed_dir = processed_dir or PROCESSED_DIR
    files = find_indicator_sheet_files(raw_dir)
    if not files:
        return None

    print("\nMeasles Indicator Sheet pipeline starting...")
    path = files[0]
    print(f"  Loading {path.name}...")
    sheet = load_indicator_sheet(path)
    summary = build_key_indicators_summary(sheet)
    print(f"    {sheet['year']} sheet: {summary['districts_covered']} districts, "
          f"{summary['total_cases_reported']} total cases reported")

    processed_dir.mkdir(parents=True, exist_ok=True)
    with open(processed_dir / "vpd_indicator_summary.json", "w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2, default=str)
    print("Measles Indicator Sheet pipeline finished OK.")
    return summary
